import openpyxl
from openpyxl.styles import PatternFill

def load_workbook(file):
    """Helper function to load the workbook."""
    try:
        return openpyxl.load_workbook(file)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file}' was not found.")
    except Exception as e:
        raise Exception(f"An error occurred while loading the workbook: {e}")

def get_worksheet(wb, sheet_name):
    """Helper function to get the worksheet."""
    try:
        return wb[sheet_name]
    except KeyError:
        raise KeyError(f"The sheet '{sheet_name}' does not exist in the workbook.")

def get_row_count(file, sheet_name):
    """Get the total number of rows in the specified sheet."""
    wb = load_workbook(file)
    ws = get_worksheet(wb, sheet_name)
    return ws.max_row

def get_column_count(file, sheet_name):
    """Get the total number of columns in the specified sheet."""
    wb = load_workbook(file)
    ws = get_worksheet(wb, sheet_name)
    return ws.max_column

def read_data(file, sheet_name, row_num, col_num):
    """Read data from a specific cell."""
    wb = load_workbook(file)
    ws = get_worksheet(wb, sheet_name)
    try:
        return ws.cell(row=row_num, column=col_num).value
    except IndexError:
        raise IndexError(f"Invalid row or column number: ({row_num}, {col_num}).")

def write_data(file, data, sheet_name, row_num, col_num):
    """Write data to a specific cell."""
    wb = load_workbook(file)
    ws = get_worksheet(wb, sheet_name)
    try:
        ws.cell(row=row_num, column=col_num).value = data
        wb.save(file)
    except IndexError:
        raise IndexError(f"Invalid row or column number: ({row_num}, {col_num}).")

def fill_cell_color(file, sheet_name, row_num, col_num, color):
    """Fill a cell with a specific color."""
    wb = load_workbook(file)
    ws = get_worksheet(wb, sheet_name)
    try:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws.cell(row=row_num, column=col_num).fill = fill
        wb.save(file)
    except IndexError:
        raise IndexError(f"Invalid row or column number: ({row_num}, {col_num}).")

# Example usage
if __name__ == "__main__":
    file = "example.xlsx"
    sheet_name = "Sheet1"

    # Get row and column count
    print(f"Rows: {get_row_count(file, sheet_name)}")
    print(f"Columns: {get_column_count(file, sheet_name)}")

    # Read data
    print(f"Data at (1, 1): {read_data(file, sheet_name, 1, 1)}")

    # Write data
    write_data(file, "Hello", sheet_name, 1, 1)

    # Fill cell with green color
    fill_cell_color(file, sheet_name, 1, 1, "60b212")  # Green

    # Fill cell with red color
    fill_cell_color(file, sheet_name, 2, 2, "ff0000")  # Red